"""
Reconciliation Report Parser

Parse reconciliation report CSV/XLSX files with support for DSL filtering and aggregation
(WHERE / SELECT / GROUP BY / ORDER BY / LIMIT).
"""

import csv
import json
import os
from decimal import Decimal, ROUND_HALF_EVEN, InvalidOperation
from pathlib import Path
from typing import List, Dict, Any, Optional, Union

from .constants import (
    ALL_FEE_FIELDS,
    SETTLEMENT_DETAIL_FILENAME_RE,
    SETTLEMENT_DETAIL_CORE_COLUMNS,
    SETTLEMENT_DETAIL_FORBIDDEN_COLUMNS,
    SETTLEMENT_DETAIL_MIN_CORE_MATCH,
)


# ============================================================
# Settlement Detail report type gate (filename + lightweight header check)
# Filename gate is mandatory; header check defends against obvious
# renames / forgeries.
# ============================================================
ALLOWED_EXTENSIONS = (".csv", ".xlsx")


class ReportTypeError(ValueError):
    """Raised by detect_report_type() when a file is rejected."""

    def __init__(self, reason: str, file_path: str, kind: str):
        super().__init__(reason)
        self.reason = reason
        self.file_path = file_path
        # kind ∈ {"extension", "filename", "content"}
        self.kind = kind


def _read_csv_header(file_path: str) -> List[str]:
    for encoding in ("utf-8", "gbk"):
        try:
            with open(file_path, "r", encoding=encoding) as f:
                reader = csv.reader(f)
                return next(reader, []) or []
        except UnicodeDecodeError:
            continue
        except StopIteration:
            return []
    return []


def _read_xlsx_header(file_path: str) -> List[str]:
    from openpyxl import load_workbook

    # Note: for some Antom-exported XLSX files, read_only=True may return
    # only the first column (the sheet's dimension attribute does not cover
    # the full row range). Use the default (non-read-only) mode and iterate
    # up to ws.max_column to guarantee a complete header row.
    wb = load_workbook(filename=file_path, data_only=True)
    try:
        ws = wb.active
        max_col = ws.max_column or 0
        if max_col <= 0:
            return []
        return [
            (str(ws.cell(row=1, column=c).value).strip()
             if ws.cell(row=1, column=c).value is not None else "")
            for c in range(1, max_col + 1)
        ]
    finally:
        wb.close()


def detect_report_type(file_path: str) -> Dict[str, Any]:
    """Validate whether the file is a Settlement Detail report.

    Returns ``{"report_type": "SETTLEMENT_DETAIL", "file_path": ...}`` on
    success; raises ``ReportTypeError`` otherwise.

    Detection strategy (cheap to expensive):
      1. Extension allowlist: only ``.csv`` / ``.xlsx``.
      2. Filename gate: must match ``SETTLEMENT_DETAIL_FILENAME_RE``.
      3. Header content check (positive + negative):
         a. Positive: header must match at least ``SETTLEMENT_DETAIL_MIN_CORE_MATCH``
            columns from ``SETTLEMENT_DETAIL_CORE_COLUMNS`` (rejects renamed Summary).
         b. Negative: header must not contain any ``SETTLEMENT_DETAIL_FORBIDDEN_COLUMNS``
            column (rejects renamed Transaction Detail).
    """
    file_name = os.path.basename(file_path)
    ext = os.path.splitext(file_name)[1].lower()

    if ext not in ALLOWED_EXTENSIONS:
        raise ReportTypeError(
            reason=(
                "Only CSV or XLSX report files are accepted. "
                "Please re-export the report in one of these two formats before submitting."
            ),
            file_path=file_path,
            kind="extension",
        )

    if not SETTLEMENT_DETAIL_FILENAME_RE.match(file_name):
        raise ReportTypeError(
            reason=(
                "Only Settlement Detail reports can be analyzed. This file does not appear to be a "
                "Settlement Detail report. Please download a Settlement Detail report from the merchant "
                "dashboard and resubmit. "
                "(The filename must contain both `SETTLEMENT` and `DETAIL`, "
                "e.g. `SETTLEMENT_DETAIL_*.xlsx` or `Settlement_Detail_*.csv`.)"
            ),
            file_path=file_path,
            kind="filename",
        )

    try:
        header = _read_xlsx_header(file_path) if ext == ".xlsx" else _read_csv_header(file_path)
    except Exception as e:
        raise ReportTypeError(
            reason=f"Failed to read report header; type detection cannot proceed: {e}",
            file_path=file_path,
            kind="content",
        )

    header_set = {col for col in header if col}

    # Positive signal: must match at least N Settlement Detail core columns
    core_matches = header_set & SETTLEMENT_DETAIL_CORE_COLUMNS
    if len(core_matches) < SETTLEMENT_DETAIL_MIN_CORE_MATCH:
        raise ReportTypeError(
            reason=(
                "Only Settlement Detail reports can be analyzed. This file's header does not conform "
                "to the Settlement Detail structure "
                f"(matched {len(core_matches)}/{len(SETTLEMENT_DETAIL_CORE_COLUMNS)} core columns; "
                f"at least {SETTLEMENT_DETAIL_MIN_CORE_MATCH} are required). "
                "Please re-download the Settlement Detail report from the merchant dashboard without renaming it."
            ),
            file_path=file_path,
            kind="content",
        )

    # Negative signal: must not contain TX-exclusive columns
    forbidden_hits = header_set & SETTLEMENT_DETAIL_FORBIDDEN_COLUMNS
    if forbidden_hits:
        raise ReportTypeError(
            reason=(
                "This file appears to be a Transaction Detail report, not a Settlement Detail report "
                f"(contains TX-exclusive columns: {sorted(forbidden_hits)}). "
                "Only Settlement Detail reports can be analyzed. "
                "Please re-download the Settlement Detail report from the merchant dashboard."
            ),
            file_path=file_path,
            kind="content",
        )

    return {"report_type": "SETTLEMENT_DETAIL", "file_path": file_path}


def get_fee_summary(data: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Return a summary of all 20 fee fields in a single pass, eliminating
    the LLM's freedom to selectively choose fields.

    Uses Decimal arithmetic for financial precision, consistent with validators.py.
    All values are constructed from strings and quantized to 2 decimal places
    using ROUND_HALF_EVEN (banker's rounding).
    """
    Q = Decimal('0.01')
    fee_details = {}
    non_zero_fields = []
    total_fees = Decimal('0')

    for field in ALL_FEE_FIELDS:
        field_sum = Decimal('0')
        for row in data:
            val = row.get(field, "")
            if val is not None:
                s = str(val).strip()
                if s and s.lower() not in ("", "null", "none"):
                    try:
                        field_sum += Decimal(s)
                    except (InvalidOperation, ValueError):
                        pass
        field_sum = field_sum.quantize(Q, rounding=ROUND_HALF_EVEN)
        fee_details[field] = field_sum
        total_fees += field_sum
        if field_sum != Decimal('0'):
            non_zero_fields.append(field)

    total_fees = total_fees.quantize(Q, rounding=ROUND_HALF_EVEN)

    # Convert Decimal to str for JSON serialization safety while preserving
    # exact decimal representation (e.g. "59.97" not 59.970000000000006).
    # Consumers that need numeric operations can Decimal(str_val) safely.
    return {
        "fee_details": {k: str(v) for k, v in fee_details.items()},
        "total_fees": str(total_fees),
        "non_zero_fields": non_zero_fields,
        "field_count": {
            "total": len(ALL_FEE_FIELDS),
            "non_zero": len(non_zero_fields),
            "zero_or_empty": len(ALL_FEE_FIELDS) - len(non_zero_fields),
        },
    }


# JSON Schema validation (optional dependency)
try:
    import jsonschema
    HAS_JSONSCHEMA = True
except ImportError:
    HAS_JSONSCHEMA = False


def get_dsl_schema() -> Dict[str, Any]:
    """Load the DSL schema definition."""
    schema_path = Path(__file__).parent / "dsl_schema.json"
    with open(schema_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def validate_dsl(filters: Dict[str, Any]) -> Optional[str]:
    """Validate a DSL object against the schema. Returns None or an error message."""
    if not filters:
        return None
    
    if not HAS_JSONSCHEMA:
        # jsonschema not installed; skip validation
        return None
    
    try:
        schema = get_dsl_schema()
        jsonschema.validate(instance=filters, schema=schema)
        return None
    except jsonschema.ValidationError as e:
        return f"DSL validation failed: {e.message}"


class FilterDSL:
    """DSL filter executor."""
    
    def __init__(self, filters: Optional[Dict[str, Any]] = None):
        """Initialize the DSL executor."""
        self.filters = filters or {}
    
    def _compare_values(self, actual: Any, op: str, expected: Any) -> bool:
        """Compare two values; supports =, !=, >, >=, <, <=, IN, NOT IN, LIKE, IS NULL, IS NOT NULL."""
        if op == "=":
            return actual == expected
        elif op == "!=":
            return actual != expected
        elif op in (">", ">=", "<", "<="):
            try:
                actual_num = float(actual)
                expected_num = float(expected)
            except (ValueError, TypeError):
                return False
            if op == ">":
                return actual_num > expected_num
            elif op == ">=":
                return actual_num >= expected_num
            elif op == "<":
                return actual_num < expected_num
            else:
                return actual_num <= expected_num
        elif op == "IN":
            return actual in expected
        elif op == "NOT IN":
            return actual not in expected
        elif op == "LIKE":
            # Support % wildcard: escape regex special chars then restore %
            import re
            escaped = re.escape(expected)
            pattern = escaped.replace("%", ".*")
            return bool(re.match(f"^{pattern}$", str(actual)))
        elif op == "IS NULL":
            return actual is None or actual == "" or str(actual).lower() == "null"
        elif op == "IS NOT NULL":
            return not (actual is None or actual == "" or str(actual).lower() == "null")
        else:
            raise ValueError(f"Unsupported operator: {op}")
    
    def _evaluate_condition(self, row: Dict[str, Any], condition: Dict[str, Any]) -> bool:
        """Evaluate a single condition (supports AND/OR nesting)."""
        # AND/OR nesting
        if "AND" in condition:
            return all(self._evaluate_condition(row, c) for c in condition["AND"])
        
        if "OR" in condition:
            return any(self._evaluate_condition(row, c) for c in condition["OR"])
        
        # Simple condition
        column = condition.get("column")
        op = condition.get("op")
        value = condition.get("value")
        
        if column is None or op is None:
            return True
        
        actual_value = row.get(column)
        return self._compare_values(actual_value, op, value)
    
    def apply_where(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Apply WHERE filtering."""
        where_clause = self.filters.get("WHERE")
        if not where_clause:
            return rows
        
        return [row for row in rows if self._evaluate_condition(row, where_clause)]
    
    def apply_select(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Apply SELECT column projection."""
        select_columns = self.filters.get("SELECT")
        if not select_columns:
            return rows
        
        return [
            {k: v for k, v in row.items() if k in select_columns}
            for row in rows
        ]
    
    def _apply_aggregation(self, rows: List[Dict[str, Any]], agg: Dict[str, Any]) -> Any:
        """Apply a single aggregation function (COUNT/SUM/AVG/MIN/MAX/FIRST/LAST).

        Uses Decimal for SUM/AVG/MIN/MAX to ensure financial precision,
        consistent with get_fee_summary() and validators.py.
        Returns str for Decimal results (JSON-safe), int for COUNT, raw values for FIRST/LAST.
        """
        Q = Decimal('0.01')
        column = agg.get("column")
        func = agg.get("function", "COUNT")

        values = [row.get(column) for row in rows]

        def _to_decimal_values(vals):
            """Safely convert non-empty values to Decimal, skipping unparseable ones
            (e.g. "null", "N/A") that would otherwise crash a generator-based sum()."""
            result = []
            for v in vals:
                if v is None or v == "":
                    continue
                try:
                    result.append(Decimal(str(v)))
                except (InvalidOperation, ValueError, TypeError):
                    continue
            return result

        if func == "COUNT":
            return len([v for v in values if v is not None and v != ""])
        elif func == "SUM":
            numeric_values = _to_decimal_values(values)
            if not numeric_values:
                return "0"
            total = sum(numeric_values)
            return str(total.quantize(Q, rounding=ROUND_HALF_EVEN))
        elif func == "AVG":
            numeric_values = _to_decimal_values(values)
            if not numeric_values:
                return "0"
            avg = sum(numeric_values) / len(numeric_values)
            return str(avg.quantize(Q, rounding=ROUND_HALF_EVEN))
        elif func == "MIN":
            numeric_values = _to_decimal_values(values)
            return str(min(numeric_values)) if numeric_values else None
        elif func == "MAX":
            numeric_values = _to_decimal_values(values)
            return str(max(numeric_values)) if numeric_values else None
        elif func == "FIRST":
            return values[0] if values else None
        elif func == "LAST":
            return values[-1] if values else None
        else:
            raise ValueError(f"Unsupported aggregation function: {func}")
    
    def apply_group_by(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Apply GROUP BY aggregation."""
        group_by_clause = self.filters.get("GROUP_BY")
        if not group_by_clause:
            return rows
        
        group_columns = group_by_clause.get("columns", [])
        aggregations = group_by_clause.get("aggregations", [])
        
        if not group_columns:
            # No grouping columns: aggregate the entire dataset
            result = {}
            for agg in aggregations:
                alias = agg.get("alias", agg.get("column"))
                result[alias] = self._apply_aggregation(rows, agg)
            return [result]
        
        # Group rows
        groups = {}
        for row in rows:
            key = tuple(row.get(col) for col in group_columns)
            if key not in groups:
                groups[key] = []
            groups[key].append(row)
        
        # Apply aggregations per group
        results = []
        for key, group_rows in groups.items():
            result = {col: val for col, val in zip(group_columns, key)}
            for agg in aggregations:
                alias = agg.get("alias", agg.get("column"))
                result[alias] = self._apply_aggregation(group_rows, agg)
            results.append(result)
        
        return results
    
    def apply_order_by(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Apply ORDER BY sorting (supports multiple columns).
        Returns a new sorted list without mutating the input."""
        order_by_clause = self.filters.get("ORDER_BY")
        if not order_by_clause:
            return rows
        
        result = list(rows)  # shallow copy to avoid mutating input
        # Multi-column sort (apply in reverse order for stable result)
        for order in reversed(order_by_clause):
            column = order.get("column")
            direction = order.get("direction", "ASC").upper()
            reverse = (direction == "DESC")
            
            try:
                result.sort(
                    key=lambda x: float(x.get(column, 0) or 0),
                    reverse=reverse
                )
            except (ValueError, TypeError):
                # Non-numeric: sort as string
                result.sort(
                    key=lambda x: str(x.get(column, "")),
                    reverse=reverse
                )
        
        return result
    
    def apply_limit(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Apply LIMIT."""
        limit = self.filters.get("LIMIT")
        if limit is None:
            return rows
        
        return rows[:limit]
    
    def execute(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Execute the full DSL pipeline (order: WHERE → GROUP BY → ORDER BY → LIMIT → SELECT)."""
        result = rows
        
        # 1. WHERE
        result = self.apply_where(result)
        
        # 2. GROUP BY
        result = self.apply_group_by(result)
        
        # 3. ORDER BY
        result = self.apply_order_by(result)
        
        # 4. LIMIT
        result = self.apply_limit(result)
        
        # 5. SELECT
        result = self.apply_select(result)
        
        return result


def detect_end_marker(row: Dict[str, Any]) -> bool:
    """Detect whether a row is an END marker row.
    A row is an END marker if any field contains '<END>' or equals 'END'."""
    for val in row.values():
        if val is not None:
            val_str = str(val).strip()
            if "<END>" in val_str or val_str.upper() == "END":
                return True
    return False


def parse_csv_file(file_path: str) -> Dict[str, Any]:
    """Parse a single CSV file; returns data + metadata (including empty batch detection)."""
    rows = []
    end_marker_rows = 0
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_dict = dict(row)
                # Detect END marker rows
                if detect_end_marker(row_dict):
                    end_marker_rows += 1
                else:
                    rows.append(row_dict)
    except UnicodeDecodeError:
        # Retry with alternative encoding
        try:
            with open(file_path, 'r', encoding='gbk') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    row_dict = dict(row)
                    if detect_end_marker(row_dict):
                        end_marker_rows += 1
                    else:
                        rows.append(row_dict)
        except Exception as e:
            return {
                "success": False,
                "data": [],
                "metadata": {
                    "file_path": file_path,
                    "error": f"Encoding error: {str(e)}"
                }
            }
    except Exception as e:
        return {
            "success": False,
            "data": [],
            "metadata": {
                "file_path": file_path,
                "error": f"Parse failed: {str(e)}"
            }
        }
    
    # Compute metadata
    total_rows = len(rows) + end_marker_rows
    is_empty_batch = (len(rows) == 0)
    
    metadata = {
        "file_path": file_path,
        "total_rows": total_rows,
        "data_rows": len(rows),
        "end_marker_rows": end_marker_rows,
        "is_empty_batch": is_empty_batch,
        "reason": None,
        "business_meaning": None
    }
    
    # Determine the reason and business meaning for an empty batch
    if is_empty_batch:
        if total_rows == 0:
            metadata["reason"] = "CSV file is completely empty (no header, no data)"
        elif total_rows == end_marker_rows:
            metadata["reason"] = "CSV contains only END markers; no data rows"
        else:
            metadata["reason"] = "CSV contains only a header row; no data rows"
        
        metadata["business_meaning"] = (
            "Empty batch: the merchant's pending settlement balance has not yet reached "
            "the agreed settlement threshold, so there is no actual settlement content today."
        )
    
    return {
        "success": True,
        "data": rows,
        "metadata": metadata
    }


def parse_xlsx_file(file_path: str) -> Dict[str, Any]:
    """Parse a single XLSX file; returns data + metadata (including empty batch detection)."""
    from openpyxl import load_workbook
    
    rows = []
    end_marker_rows = 0
    
    try:
        # Fix: read_only=True cannot be used as it may produce incomplete headers
        # for some files
        wb = load_workbook(filename=file_path, read_only=False)
        ws = wb.active
        
        # Read header row
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value) if cell.value is not None else "")
        
        # Read data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_dict = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    value = cell.value
                    # Normalize to string (consistent with CSV).
                    # For numeric values, route through Decimal to avoid
                    # float→str precision loss (e.g. str(float(19.99)) → "19.990000000000002").
                    if value is None:
                        value = ""
                    elif isinstance(value, float):
                        value = str(Decimal(str(value)))
                    elif isinstance(value, int):
                        value = str(value)
                    else:
                        value = str(value)
                    row_dict[headers[col_idx]] = value
            
            # Detect END marker
            if detect_end_marker(row_dict):
                end_marker_rows += 1
            else:
                rows.append(row_dict)
        
        wb.close()
    
    except Exception as e:
        return {
            "success": False,
            "data": [],
            "metadata": {
                "file_path": file_path,
                "error": f"XLSX parse failed: {str(e)}"
            }
        }
    
    # Compute metadata
    total_rows = len(rows) + end_marker_rows
    is_empty_batch = (len(rows) == 0)
    
    metadata = {
        "file_path": file_path,
        "total_rows": total_rows,
        "data_rows": len(rows),
        "end_marker_rows": end_marker_rows,
        "is_empty_batch": is_empty_batch,
        "reason": "No data rows in XLSX file" if is_empty_batch else None,
        "business_meaning": (
            "Empty batch: the merchant's pending settlement balance has not yet reached "
            "the agreed settlement threshold, so there is no actual settlement content today."
        ) if is_empty_batch else None
    }
    
    return {
        "success": True,
        "data": rows,
        "metadata": metadata
    }


def parse_reports(
    input: Union[List[str], str],
    filters: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """Main report parsing function: parse CSV/XLSX files, apply DSL filters,
    return data + fee_summary + metadata."""
    # Normalize input
    if isinstance(input, str):
        input = [input]
    
    if not isinstance(input, list):
        raise ValueError("input must be a list of strings or a single string")

    # ── Report type gate (filename + lightweight header check) ──
    # Must run before any parsing. If any file fails, the entire batch fails —
    # the agent surfaces ReportTypeError.reason to the merchant.
    for file_path in input:
        detect_report_type(file_path)

    # Parse all files
    all_rows = []
    parsed_files = []  # Successfully parsed file paths
    parse_errors = []  # Files that failed to parse
    
    file_metadata = {
        "file_count": len(input),
        "success_files": 0,
        "failed_files": 0,
        "total_rows": 0,
        "data_rows": 0,
        "end_marker_rows": 0,
        "empty_batch_files": [],
        "parse_errors": []
    }
    
    for file_path in input:
        # Select parser based on file extension
        if file_path.lower().endswith('.xlsx'):
            result = parse_xlsx_file(file_path)
        elif file_path.lower().endswith('.csv'):
            result = parse_csv_file(file_path)
        else:
            # Default to CSV
            result = parse_csv_file(file_path)
        
        if not result.get("success"):
            # Parse failure: record and continue with remaining files
            parse_errors.append({
                "file_path": file_path,
                "error": result.get("metadata", {}).get("error", "Unknown parse error")
            })
            file_metadata["failed_files"] += 1
            continue
        
        # Success: accumulate data rows
        parsed_files.append(file_path)
        all_rows.extend(result.get("data", []))
        
        # Accumulate metadata
        meta = result.get("metadata", {})
        file_metadata["total_rows"] += meta.get("total_rows", 0)
        file_metadata["data_rows"] += meta.get("data_rows", 0)
        file_metadata["end_marker_rows"] += meta.get("end_marker_rows", 0)
        
        # Record empty batch files
        if meta.get("is_empty_batch"):
            file_metadata["empty_batch_files"].append(file_path)
    
    # Update success file count
    file_metadata["success_files"] = len(parsed_files)
    file_metadata["parse_errors"] = parse_errors
    
    # Determine whether all successfully parsed files are empty batches
    is_all_empty = (len(all_rows) == 0) and (len(file_metadata["empty_batch_files"]) == file_metadata["success_files"])
    
    # Validate DSL filters before execution
    if filters:
        error_msg = validate_dsl(filters)
        if error_msg:
            return {
                "success": False,
                "partial_success": len(parsed_files) > 0,
                "data": [],
                "metadata": {
                    **file_metadata,
                    "error": error_msg
                }
            }
    
    # Apply DSL filtering (only when data rows exist)
    if filters and all_rows:
        dsl = FilterDSL(filters)
        all_rows = dsl.execute(all_rows)
    
    # Build return payload
    partial_success = len(parse_errors) > 0 and len(parsed_files) > 0
    has_success = len(parsed_files) > 0
    
    # Auto-compute fee summary (full coverage of all 20 fee fields)
    fee_summary = get_fee_summary(all_rows) if all_rows else None
    
    result = {
        "success": has_success,
        "partial_success": partial_success,
        "data": all_rows,
        "fee_summary": fee_summary,
        "metadata": file_metadata
    }
    
    # Annotate empty batch with business meaning
    if is_all_empty and has_success:
        result["metadata"]["is_empty_batch"] = True
        result["metadata"]["business_meaning"] = (
            "Empty batch: the merchant's pending settlement balance has not yet reached "
            "the agreed settlement threshold, so there is no actual settlement content. "
            "This is not an error — it is a normal business state."
        )
    
    return result


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import argparse
    import sys

    _cli = argparse.ArgumentParser(
        description="Parse settlement detail CSV/XLSX reports",
    )
    _cli.add_argument("--files", nargs="+", required=True,
                      help="CSV/XLSX file paths to parse")
    _cli.add_argument("--filters", default=None,
                      help="DSL filter JSON string (optional)")
    _args = _cli.parse_args()

    _filters = json.loads(_args.filters) if _args.filters else None
    _result = parse_reports(_args.files, _filters)
    print(json.dumps(_result, indent=2, ensure_ascii=False, default=str))
    sys.exit(0 if _result.get("success") else 1)